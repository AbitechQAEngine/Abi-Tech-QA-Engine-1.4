from fastapi import APIRouter, UploadFile, File, HTTPException, Depends, Form
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional, List
from groq import Groq
from openai import OpenAI
import asyncio
import os, io, json
import pandas as pd
from openpyxl import Workbook
from sqlalchemy.orm import Session

from database import get_db
import models
from deps import get_current_user, get_owned_project

router = APIRouter()

def get_client():
    api_key = os.getenv("GROQ_API_KEY")
    if not api_key:
        raise HTTPException(status_code=500, detail="GROQ_API_KEY not set")
    return Groq(api_key=api_key)


def get_openrouter_client():
    """Client for Llama 4 Scout via OpenRouter (replaces Groq for this model,
    since Groq stopped reliably serving llama-4-scout)."""
    api_key = os.getenv("OPENROUTER_API_KEY")
    if not api_key:
        raise HTTPException(status_code=500, detail="OPENROUTER_API_KEY not set")
    return OpenAI(
        api_key=api_key,
        base_url="https://openrouter.ai/api/v1",
    )


def _escape_raw_control_chars_in_strings(text: str) -> str:
    """Models very often emit a literal newline/tab inside a JSON string
    value (e.g. inside "steps") instead of the escaped \\n. That is invalid
    JSON and is the most common cause of 'Expecting , delimiter' errors.
    This walks the text once, tracking whether we're inside a string
    literal, and escapes raw control characters found there."""
    out = []
    in_string = False
    escaped = False
    for ch in text:
        if in_string:
            if escaped:
                out.append(ch)
                escaped = False
                continue
            if ch == "\\":
                out.append(ch)
                escaped = True
                continue
            if ch == '"':
                out.append(ch)
                in_string = False
                continue
            if ch == "\n":
                out.append("\\n")
                continue
            if ch == "\r":
                continue
            if ch == "\t":
                out.append("\\t")
                continue
            out.append(ch)
        else:
            if ch == '"':
                in_string = True
            out.append(ch)
    return "".join(out)


def _recover_truncated_array(text: str):
    """If the model's response got cut off mid-array (hit max_tokens), try
    to salvage the complete objects that were returned and close the array."""
    last_complete = text.rfind("}")
    if last_complete == -1:
        return None
    salvaged = text[: last_complete + 1] + "]"
    try:
        return json.loads(salvaged)
    except json.JSONDecodeError:
        return None


def parse_test_cases_json(raw: str):
    """Parse the AI's response into a JSON array of test cases, tolerating
    the common formatting slips models make (code fences, stray prose
    before/after the array, trailing commas, raw newlines inside strings,
    truncated output)."""
    import re

    content = raw.strip()

    # Strip ```json ... ``` or ``` ... ``` code fences
    if "```" in content:
        parts = content.split("```")
        # Prefer a fenced block that actually contains a JSON array
        candidates = [p[4:] if p.strip().lower().startswith("json") else p for p in parts]
        candidates = [c.strip() for c in candidates if "[" in c]
        if candidates:
            content = candidates[0]

    # If there's extra prose around the array, isolate the outermost [ ... ]
    start = content.find("[")
    if start != -1:
        content = content[start:]
    end = content.rfind("]")
    if end != -1:
        content = content[: end + 1]

    def try_load(text):
        return json.loads(text)

    attempts = [content]

    # Fixup 1: escape raw control characters that landed inside string values
    attempts.append(_escape_raw_control_chars_in_strings(content))

    # Fixup 2: trailing commas before a closing ] or }
    attempts.append(re.sub(r",\s*([\]}])", r"\1", attempts[-1]))

    # Fixup 3: missing comma between adjacent objects "}\s*{"
    attempts.append(re.sub(r"}\s*{", "},{", attempts[-1]))

    last_error = None
    for candidate in attempts:
        try:
            return try_load(candidate)
        except json.JSONDecodeError as e:
            last_error = e

    # Last resort: the response may have been cut off before it finished —
    # salvage whatever complete test case objects are present.
    recovered = _recover_truncated_array(attempts[-1])
    if recovered:
        return recovered

    raise HTTPException(
        status_code=502,
        detail=(
            "The AI response could not be parsed as valid test case data "
            f"({last_error.msg} at line {last_error.lineno}, column {last_error.colno}). "
            "Please try generating again."
        ),
    )

def smart_rename(df):
    """Rename columns to standard names using lowercase + strip matching."""
    # First strip ALL column names of whitespace and special chars
    df.columns = [str(c).strip().replace('\xa0', ' ').replace('\t', ' ') for c in df.columns]

    print("DEBUG columns after strip:", list(df.columns))
    print("DEBUG columns repr:", [repr(c) for c in df.columns])

    rename = {}
    for col in df.columns:
        low = col.lower().strip()
        print(f"  checking: {repr(col)} → lower: {repr(low)}")

        if low in ["title", "test title", "test case title", "test name",
                   "name", "scenario", "description", "test case name"]:
            rename[col] = "title"

        elif low in ["steps", "test steps", "step", "steps to reproduce",
                     "execution steps", "actions", "procedure", "test procedure",
                     "test action", "test actions"]:
            rename[col] = "steps"

        elif low in ["expected", "expected result", "expected results",
                     "expected outcome", "expected output", "expected behavior",
                     "expected behaviour", "result", "outcome",
                     "pass criteria", "acceptance criteria", "expected value"]:
            rename[col] = "expected"

        elif low in ["id", "test id", "tc id", "test case id",
                     "sl no", "sl.no", "s.no", "sno", "no", "sr no", "sr.no"]:
            rename[col] = "id"

        elif low in ["type", "test type", "case type", "test case type"]:
            rename[col] = "type"

        elif low in ["priority", "test priority", "case priority"]:
            rename[col] = "priority"

    print("DEBUG rename map:", rename)
    df = df.rename(columns=rename)
    print("DEBUG columns after rename:", list(df.columns))
    return df


class TestCaseRequest(BaseModel):
    project_id: int
    module: str
    feature: str
    user_story: Optional[str] = ""
    test_types: List[str] = ["positive", "negative", "validation", "boundary"]
    # If not provided, generation continues from project.test_case_counter + 1.
    # If provided, generation starts from this number instead (BRD Module 7, "No" path).
    custom_start_id: Optional[int] = None

SYSTEM_PROMPT = """You are an expert QA engineer. Generate ALL POSSIBLE test cases based on the
provided information — do not limit the count or hold back coverage. Be exhaustive: cover every
positive, negative, validation, boundary, edge-case, and error-handling scenario you can identify
for the given module/feature, including uncommon and edge conditions a thorough QA engineer would
think of. Do not skip a scenario just to keep the list short.
Return ONLY a valid JSON array with no markdown formatting, no code blocks, just raw JSON.
Each object must have: id, title, type, steps, expected, priority.
- id: sequential like TC_001, TC_002
- title: concise test case name
- type: positive/negative/validation/boundary
- steps: numbered steps as a single string separated by \\n
- expected: expected result
- priority: High/Medium/Low"""

@router.get("/continue-check/{project_id}")
async def check_continue(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """BRD Module 7: before generating, tell the frontend the current counter
    so it can show the 'Continue from TC-0XX?' Yes/No prompt."""
    project = get_owned_project(project_id, db, current_user)
    return {
        "current_counter": project.test_case_counter,
        "next_id_if_yes": project.test_case_counter + 1,
    }


@router.post("/generate")
async def generate_test_cases(
    request: TestCaseRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    project = get_owned_project(request.project_id, db, current_user)

    types_str = ", ".join(request.test_types)
    prompt = f"""Generate ALL POSSIBLE test cases for:
Module: {request.module}
Feature: {request.feature}
User Story: {request.user_story}
Test Types needed: {types_str}
Do not limit yourself to a fixed number per type -- generate every relevant positive, negative,
validation, and boundary test case you can think of for this feature, covering all realistic
scenarios, edge cases, and error conditions. Be exhaustive rather than brief.
Return as a JSON array only."""
    try:
        client = get_openrouter_client()
        response = await asyncio.to_thread(
            client.chat.completions.create,
            model="meta-llama/llama-3.3-70b-instruct",
            messages=[
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user", "content": prompt}
            ],
            temperature=0.3,
            max_tokens=8000,
            extra_headers={
                "HTTP-Referer": os.getenv("APP_URL", "http://localhost"),
                "X-Title": "Abitech QA Engine",
            },
        )
        content = response.choices[0].message.content.strip()
        test_cases = parse_test_cases_json(content)
        if request.custom_start_id is not None:
            if request.custom_start_id <= 0:
                raise HTTPException(status_code=422, detail="Starting Test Case ID must be greater than zero")
            start_id = request.custom_start_id
        else:
            start_id = project.test_case_counter + 1

        for i, tc in enumerate(test_cases):
            tc["id"] = f"TC-{str(start_id + i).zfill(3)}"

        end_id = start_id + len(test_cases) - 1

        history = models.TestCaseHistory(
            project_id=project.id,
            prompt=prompt,
            generated_test_cases=json.dumps(test_cases),
            starting_test_case_id=start_id,
            ending_test_case_id=end_id,
        )
        db.add(history)

        # Persist each test case as its own row so the project's test case
        # table can be viewed, edited and deleted, and so future generations
        # keep appending from the correct sequence number.
        for i, tc in enumerate(test_cases):
            row = models.TestCase(
                project_id=project.id,
                tc_number=start_id + i,
                display_id=tc["id"],
                title=tc.get("title", ""),
                type=tc.get("type", ""),
                steps=tc.get("steps", ""),
                expected=tc.get("expected", ""),
                priority=tc.get("priority", "Medium"),
                source="generated",
            )
            db.add(row)

        project.test_case_counter = max(project.test_case_counter, end_id)
        db.commit()

        return {
            "test_cases": test_cases,
            "count": len(test_cases),
            "starting_test_case_id": start_id,
            "ending_test_case_id": end_id,
            "project_test_case_counter": project.test_case_counter,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/history/{project_id}")
async def get_test_case_history(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    project = get_owned_project(project_id, db, current_user)
    rows = (
        db.query(models.TestCaseHistory)
        .filter(models.TestCaseHistory.project_id == project.id)
        .order_by(models.TestCaseHistory.created_at.desc())
        .all()
    )
    return [
        {
            "id": r.id,
            "prompt": r.prompt,
            "test_cases": json.loads(r.generated_test_cases),
            "starting_test_case_id": r.starting_test_case_id,
            "ending_test_case_id": r.ending_test_case_id,
            "created_at": r.created_at,
        }
        for r in rows
    ]


@router.get("/project/{project_id}")
async def list_project_test_cases(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """All persisted test cases for a project, in sequence order. Used to
    show already-generated test cases when a user re-opens a project."""
    project = get_owned_project(project_id, db, current_user)
    rows = (
        db.query(models.TestCase)
        .filter(models.TestCase.project_id == project.id)
        .order_by(models.TestCase.tc_number.asc())
        .all()
    )
    return [
        {
            "db_id": r.id,
            "id": r.display_id,
            "tc_number": r.tc_number,
            "title": r.title,
            "type": r.type,
            "steps": r.steps,
            "expected": r.expected,
            "priority": r.priority,
            "source": r.source,
            "created_at": r.created_at,
        }
        for r in rows
    ]


@router.put("/{test_case_id:int}")
async def update_test_case(
    test_case_id: int,
    payload: dict,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    row = db.query(models.TestCase).filter(models.TestCase.id == test_case_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Test case not found")
    get_owned_project(row.project_id, db, current_user)  # ownership check

    for field in ("title", "type", "steps", "expected", "priority"):
        if field in payload and payload[field] is not None:
            setattr(row, field, payload[field])
    db.commit()
    db.refresh(row)
    return {
        "db_id": row.id,
        "id": row.display_id,
        "tc_number": row.tc_number,
        "title": row.title,
        "type": row.type,
        "steps": row.steps,
        "expected": row.expected,
        "priority": row.priority,
        "source": row.source,
        "created_at": row.created_at,
    }


@router.delete("/{test_case_id:int}", status_code=204)
async def delete_test_case(
    test_case_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    row = db.query(models.TestCase).filter(models.TestCase.id == test_case_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Test case not found")
    get_owned_project(row.project_id, db, current_user)  # ownership check
    db.delete(row)
    db.commit()
    return None


@router.post("/export/excel")
async def export_excel(test_cases: List[dict]):
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"
    headers = ["ID", "Title", "Type", "Steps", "Expected Result", "Priority"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    for row, tc in enumerate(test_cases, 2):
        ws.cell(row=row, column=1, value=tc.get("id", ""))
        ws.cell(row=row, column=2, value=tc.get("title", ""))
        ws.cell(row=row, column=3, value=tc.get("type", ""))
        ws.cell(row=row, column=4, value=tc.get("steps", ""))
        ws.cell(row=row, column=5, value=tc.get("expected", ""))
        ws.cell(row=row, column=6, value=tc.get("priority", ""))
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=test_cases.xlsx"})


@router.post("/export/csv")
async def export_csv(test_cases: List[dict]):
    df = pd.DataFrame(test_cases)
    stream = io.StringIO()
    df.to_csv(stream, index=False)
    stream.seek(0)
    return StreamingResponse(iter([stream.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=test_cases.csv"})


@router.post("/upload")
async def upload_test_cases(file: UploadFile = File(...)):
    if not file.filename.endswith((".xlsx", ".csv")):
        raise HTTPException(status_code=400, detail="Only .xlsx and .csv files are supported")
    contents = await file.read()
    try:
        if file.filename.endswith(".xlsx"):
            # Read Excel — try multiple engines
            try:
                df = pd.read_excel(io.BytesIO(contents), engine="openpyxl")
            except Exception:
                df = pd.read_excel(io.BytesIO(contents))
        else:
            try:
                df = pd.read_csv(io.StringIO(contents.decode("utf-8")))
            except Exception:
                df = pd.read_csv(io.StringIO(contents.decode("latin-1")))

        # Drop fully empty rows and columns
        df = df.dropna(how="all")
        df = df.dropna(axis=1, how="all")

        # Apply smart rename
        df = smart_rename(df)

        # Check required columns
        missing = {"title", "steps", "expected"} - set(df.columns)
        if missing:
            raise HTTPException(
                status_code=422,
                detail=f"Missing columns: {missing}. Found in file: {list(df.columns)}"
            )

        test_cases = df.to_dict(orient="records")
        for i, tc in enumerate(test_cases):
            # Clean NaN
            for key in list(tc.keys()):
                try:
                    if pd.isna(tc[key]):
                        tc[key] = ""
                except Exception:
                    pass
            # Ensure required fields are strings
            tc["id"]       = str(tc.get("id", f"TC_{str(i+1).zfill(3)}")).strip()
            tc["title"]    = str(tc.get("title", "")).strip()
            tc["steps"]    = str(tc.get("steps", "")).strip()
            tc["expected"] = str(tc.get("expected", "")).strip()
            tc.setdefault("type", "positive")
            tc.setdefault("priority", "Medium")

        return {"test_cases": test_cases, "count": len(test_cases)}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"File parse error: {str(e)}")


@router.post("/debug-upload")
async def debug_upload(file: UploadFile = File(...)):
    """Debug: see exact column names Python reads from your file."""
    contents = await file.read()
    try:
        if file.filename.endswith(".xlsx"):
            df = pd.read_excel(io.BytesIO(contents), engine="openpyxl")
        else:
            df = pd.read_csv(io.StringIO(contents.decode("utf-8")))
        return {
            "columns_raw":   list(df.columns),
            "columns_lower": [str(c).lower().strip() for c in df.columns],
            "columns_repr":  [repr(c) for c in df.columns],
            "row_count":     len(df),
            "first_row":     df.iloc[0].to_dict() if len(df) > 0 else {}
        }
    except Exception as e:
        return {"error": str(e)}


@router.post("/from-screenshot")
async def generate_from_screenshot(
    files: List[UploadFile] = File(...),
    project_id: int = Form(...),
    user_story: Optional[str] = Form(None),
    description: Optional[str] = Form(None),
    test_type: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """Generate test cases from one or more UI screenshots using vision,
    combined with optional User Story / Description / Test Type context
    (Module 2: Screenshot Test Case Enhancement)."""
    project = get_owned_project(project_id, db, current_user)

    if not files:
        raise HTTPException(status_code=400, detail="Please upload at least one screenshot")
    if len(files) > 10:
        raise HTTPException(status_code=400, detail="You can upload a maximum of 10 images")

    import base64

    image_blocks = []
    filenames = []
    for f in files:
        if not f.filename.lower().endswith((".png", ".jpg", ".jpeg", ".webp")):
            raise HTTPException(status_code=400, detail="Only image files are supported (PNG, JPG, JPEG, WEBP)")
        contents = await f.read()
        if len(contents) > 10 * 1024 * 1024:
            raise HTTPException(status_code=400, detail=f"{f.filename} exceeds the 10 MB size limit")
        b64_image = base64.b64encode(contents).decode("utf-8")
        ext = f.filename.split(".")[-1].lower()
        mime = {"png": "image/png", "jpg": "image/jpeg", "jpeg": "image/jpeg", "webp": "image/webp"}.get(ext, "image/png")
        image_blocks.append({"type": "image_url", "image_url": {"url": f"data:{mime};base64,{b64_image}"}})
        filenames.append(f.filename)

    context_lines = []
    if user_story:
        context_lines.append(f"User Story:\n{user_story}")
    if description:
        context_lines.append(f"Additional Description / Business Context:\n{description}")
    if test_type:
        context_lines.append(f"Focus Test Type: {test_type}")
    context_block = "\n\n".join(context_lines)

    prompt = f"""You are an expert QA engineer. Analyze the {len(image_blocks)} attached UI screenshot(s) together as one connected user flow and generate ALL POSSIBLE, comprehensive, consolidated test cases -- do not limit the count or hold back coverage.

{context_block if context_block else "No additional context was provided beyond the screenshots."}

Look at every UI element visible across all screenshots (buttons, forms, inputs, dropdowns, links, labels, error states, etc.)
and generate test cases covering, exhaustively:
- Positive test cases (happy path)
- Negative test cases (invalid inputs, errors)
- Validation test cases (field validations)
- Boundary test cases (edge cases)
- Any other realistic scenario, error condition, or edge case a thorough QA engineer would test
{f"- Prioritize {test_type} style test cases where relevant" if test_type else ""}

Return ONLY a valid JSON array with no markdown, no code blocks.
Each object must have: id, title, type, steps, expected, priority.
- id: TC_001, TC_002, etc.
- title: concise test case name
- type: positive/negative/validation/boundary
- steps: numbered steps as single string separated by \\n
- expected: expected result
- priority: High/Medium/Low

Do not cap the number of test cases -- generate every relevant test case you can identify across all screenshots and the provided context, however many that is."""

    try:
        client = get_openrouter_client()

        # Use Llama 4 Scout via OpenRouter (vision-capable).
        # Run the blocking SDK call in a worker thread so it doesn't freeze
        # the event loop for the duration of the request (important with
        # WEB_CONCURRENCY=1 -- otherwise every other request, including
        # Render's health check, stalls until this one finishes).
        response = await asyncio.to_thread(
            client.chat.completions.create,
            model="meta-llama/llama-4-scout",
            messages=[
                {
                    "role": "user",
                    "content": [*image_blocks, {"type": "text", "text": prompt}]
                }
            ],
            temperature=0.3,
            max_tokens=8000,
            extra_headers={
                # Optional but recommended by OpenRouter for attribution/rankings
                "HTTP-Referer": os.getenv("APP_URL", "http://localhost"),
                "X-Title": "Abitech QA Engine",
            },
        )

        content = response.choices[0].message.content.strip()
        test_cases = parse_test_cases_json(content)
        # by the text-based generator (BRD Module 7), so screenshot-derived
        # cases slot in right after the last one shown in the table.
        start_id = project.test_case_counter + 1
        for i, tc in enumerate(test_cases):
            tc["id"] = f"TC-{str(start_id + i).zfill(3)}"
        end_id = start_id + len(test_cases) - 1

        history = models.ScreenshotAnalysisHistory(
            project_id=project.id,
            image_path=",".join(filenames),
            user_story=user_story,
            description=description,
            test_type=test_type,
            generated_analysis=json.dumps(test_cases),
        )
        db.add(history)

        for i, tc in enumerate(test_cases):
            row = models.TestCase(
                project_id=project.id,
                tc_number=start_id + i,
                display_id=tc["id"],
                title=tc.get("title", ""),
                type=tc.get("type", ""),
                steps=tc.get("steps", ""),
                expected=tc.get("expected", ""),
                priority=tc.get("priority", "Medium"),
                source="screenshot",
            )
            db.add(row)

        project.test_case_counter = max(project.test_case_counter, end_id)
        db.commit()

        return {
            "test_cases": test_cases,
            "count": len(test_cases),
            "starting_test_case_id": start_id,
            "ending_test_case_id": end_id,
            "project_test_case_counter": project.test_case_counter,
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
